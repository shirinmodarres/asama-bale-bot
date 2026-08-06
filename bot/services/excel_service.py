from datetime import datetime
from pathlib import Path
from tempfile import NamedTemporaryFile

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from bot.data.statuses import order_status_label, status_label
from bot.utils.datetime_format import format_shamsi_datetime
from data.static_data import get_sales_experts, get_store


def _unit_status_label(status: str) -> str:
    return {
        "pending": "در انتظار بررسی",
        "approved": "تایید شده",
        "rejected": "رد شده",
    }.get(status, status)


def _media_export_value(media: dict) -> str:
    if not media:
        return ""
    if media.get("type") == "photo":
        return media.get("local_path") or media.get("file_id") or "عکس ارسال شده"
    return media.get("value") or ""


def _first_unit_value(order: dict, field_name: str):
    for unit in order.get("units", []):
        value = unit.get(field_name)
        if value not in (None, ""):
            return value
    return ""


def _order_or_first_unit_value(order: dict, field_name: str):
    value = order.get(field_name)
    if value not in (None, ""):
        return value
    return _first_unit_value(order, field_name)


def _percent_export_value(value) -> str:
    if value in (None, ""):
        return ""
    text = str(value).strip()
    return text if text.endswith("٪") or text.endswith("%") else f"{text}٪"


class ExcelService:
    HEADERS = [
        "Request ID",
        "Store Code",
        "Store Name",
        "Seller Name",
        "Seller Phone",
        "Items",
        "Total Cartons",
        "Expert Name",
        "Status",
        "Expert Reject Reason",
        "Manager Reject Reason",
        "Expert Final Decision At",
        "Manager Final Decision At",
        "Created At",
        "Updated At",
    ]

    # هدرهای فارسی نمایشی برای گزارش درخواست‌ها (ترتیب دقیقاً مطابق HEADERS/داده‌های append شده)
    REQUEST_HEADERS_FA = [
        "شناسه درخواست",
        "کد فروشگاه",
        "نام فروشگاه",
        "نام فروشنده",
        "شماره تماس فروشنده",
        "اقلام",
        "مجموع کارتن",
        "نام کارشناس",
        "وضعیت",
        "دلیل رد کارشناس",
        "دلیل رد مدیر",
        "تاریخ تصمیم نهایی کارشناس",
        "تاریخ تصمیم نهایی مدیر",
        "تاریخ ایجاد",
        "تاریخ به‌روزرسانی",
    ]

    # هدرهای فارسی نمایشی برای گزارش سفارش‌ها
    ORDER_HEADERS_FA = [
        "شناسه سفارش",
        "کد فروشگاه",
        "نام فروشگاه",
        "نام فروشنده",
        "شماره تماس فروشنده",
        "نام کارشناس",
        "دسته‌بندی",
        "محصول",
        "مدل",
        "تعداد",
        "واحدها",
        "وضعیت",
        "دلیل رد",
        "تاریخ ایجاد",
        "قیمت واحد کالا (ریال)",
        "درصد پورسانت",
        "مبلغ پورسانت دریافت‌شده (ریال)",
    ]

    # ستون‌هایی که باید به‌صورت عدد صحیح (بدون اعشار) نمایش داده شوند (۱-بیس، بر اساس شماره ستون در اکسل)
    REQUEST_INTEGER_COLUMNS = {1, 7}  # شناسه درخواست، مجموع کارتن
    ORDER_INTEGER_COLUMNS = {1, 10, 15, 17}  # شناسه سفارش، تعداد، قیمت واحد، مبلغ پورسانت

    # ستون‌هایی که متن چندخطی دارند و باید wrap شوند
    REQUEST_WRAP_COLUMNS = {6}  # اقلام
    ORDER_WRAP_COLUMNS = {11}  # واحدها

    HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    HEADER_FONT = Font(name="Tahoma", size=11, bold=True, color="FFFFFF")
    BODY_FONT = Font(name="Tahoma", size=10)
    THIN_SIDE = Side(style="thin", color="B7B7B7")
    CELL_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

    def export_requests(self, requests: list[dict]) -> Path:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "درخواست‌ها"
        sheet.append(self.REQUEST_HEADERS_FA)

        for request in requests:
            store = get_store(request["store_code"]) or {}
            expert = get_sales_experts().get(store.get("expert_key"), {})
            items_text = "\n".join(
                f"{item['category_name']} / {item.get('product_name', '')} / مدل {item['product_model']} / ({item['carton_quantity']} کارتن)"
                for item in request["items"]
            )
            total_cartons = sum(item["carton_quantity"] for item in request["items"])
            sheet.append([
                request["id"],
                request["store_code"],
                request.get("store_name") or store.get("name", ""),
                request["seller_full_name"],
                request["seller_phone"],
                items_text,
                total_cartons,
                request.get("expert_name") or expert.get("full_name", ""),
                status_label(request["status"]),
                request.get("expert_reject_reason", ""),
                request.get("manager_reject_reason", ""),
                format_shamsi_datetime(request.get("expert_decision_at", "")),
                format_shamsi_datetime(request.get("manager_decision_at", "")),
                format_shamsi_datetime(request["created_at"]),
                format_shamsi_datetime(request["updated_at"]),
            ])

        self._apply_professional_style(
            sheet,
            integer_columns=self.REQUEST_INTEGER_COLUMNS,
            wrap_columns=self.REQUEST_WRAP_COLUMNS,
        )

        return self._save_workbook(workbook, prefix="sales-requests")

    def export_orders(self, orders: list[dict]) -> Path:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "سفارش‌ها"
        sheet.append(self.ORDER_HEADERS_FA)

        for order in orders:
            units_text = "\n".join(
                (
                    f"{unit['index']}. "
                    f"کد رهگیری: {_media_export_value(unit.get('tracking_code', {}))} | "
                    f"فاکتور: {_media_export_value(unit.get('factor_image', {}))} | "
                    f"وضعیت کالا: {_unit_status_label(unit.get('validation_status', 'pending'))} | "
                    f"علت رد: {unit.get('rejection_reason_text') or ''} | "
                    f"تاریخ تصمیم: {format_shamsi_datetime(unit.get('validation_decision_at', ''))}"
                )
                for unit in order.get("units", [])
            )
            sheet.append([
                order["id"],
                order["store_code"],
                order.get("store_name", ""),
                order.get("seller_name", ""),
                order.get("seller_phone", ""),
                order.get("expert_name", ""),
                order.get("category_name", ""),
                order.get("product_name", ""),
                order.get("product_model", ""),
                order.get("quantity", ""),
                units_text,
                order_status_label(order.get("status", "")),
                order.get("rejection_reason_text") or "",
                format_shamsi_datetime(order.get("created_at", "")),
                _order_or_first_unit_value(order, "product_price"),
                _percent_export_value(_order_or_first_unit_value(order, "commission_percent")),
                _order_or_first_unit_value(order, "commission_amount"),
            ])

        self._apply_professional_style(
            sheet,
            integer_columns=self.ORDER_INTEGER_COLUMNS,
            wrap_columns=self.ORDER_WRAP_COLUMNS,
        )

        return self._save_workbook(workbook, prefix="sales-orders")

    def _apply_professional_style(
        self,
        sheet: Worksheet,
        integer_columns: set[int],
        wrap_columns: set[int],
    ) -> None:
        """استایل حرفه‌ای مشترک: راست‌چین، فونت، بردر، فیلتر، فریز، فرمت عدد و عرض ستون."""

        max_col = sheet.max_column
        max_row = sheet.max_row

        # جهت راست‌چین کل شیت (مناسب محتوای فارسی)
        sheet.sheet_view.rightToLeft = True

        # استایل هدر (ردیف اول)
        for col_idx in range(1, max_col + 1):
            cell = sheet.cell(row=1, column=col_idx)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = self.CELL_BORDER
        sheet.row_dimensions[1].height = 26

        # استایل بدنه‌ی جدول
        for row_idx in range(2, max_row + 1):
            max_lines_in_row = 1
            for col_idx in range(1, max_col + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell.font = self.BODY_FONT
                cell.border = self.CELL_BORDER

                should_wrap = col_idx in wrap_columns
                cell.alignment = Alignment(
                    horizontal="right",
                    vertical="center",
                    wrap_text=should_wrap,
                )

                if col_idx in integer_columns:
                    cell.number_format = "#,##0"
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                if should_wrap and cell.value:
                    max_lines_in_row = max(max_lines_in_row, str(cell.value).count("\n") + 1)

            if max_lines_in_row > 1:
                sheet.row_dimensions[row_idx].height = min(15 * max_lines_in_row, 300)

        # عرض ستون‌ها متناسب با طولانی‌ترین محتوای هر ستون (با سقف منطقی برای خوانایی)
        for col_idx in range(1, max_col + 1):
            column_letter = get_column_letter(col_idx)
            lengths = []
            for row_idx in range(1, max_row + 1):
                value = sheet.cell(row=row_idx, column=col_idx).value
                if value is None:
                    continue
                text = str(value)
                # طولانی‌ترین خط را در نظر می‌گیریم، نه کل متن چندخطی
                longest_line = max((len(line) for line in text.split("\n")), default=0)
                lengths.append(longest_line)
            width = (max(lengths) if lengths else 10) + 4
            sheet.column_dimensions[column_letter].width = min(max(width, 12), 60)

        # فعال‌سازی فیلتر روی کل محدوده‌ی جدول
        last_col_letter = get_column_letter(max_col)
        sheet.auto_filter.ref = f"A1:{last_col_letter}{max_row}"

        # فریز کردن ردیف هدر
        sheet.freeze_panes = "A2"

    @staticmethod
    def _save_workbook(workbook: Workbook, prefix: str) -> Path:
        timestamp = datetime.now().strftime("%Y%m%d-%H%M")
        tmp = NamedTemporaryFile(
            delete=False,
            prefix=f"{prefix}-{timestamp}-",
            suffix=".xlsx",
        )
        tmp.close()
        workbook.save(tmp.name)
        return Path(tmp.name)
